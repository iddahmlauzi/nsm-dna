import re
from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

from dms.shared import (
    VariantRecord,
    describe_coding_edit,
    read_fasta,
    replace_codon,
    translate_dna,
    write_variants,
)

ASSAY_ID = "CCDB_ECOLI_Adkar_2012"
MUTATION_PATTERN = re.compile(r"([A-Z*])(\d{3})([A-Z*])=([ACGT]{3})")


def iter_variants(source_dir: Path) -> Iterator[VariantRecord]:
    """Convert the reported CcdB codon variants and raw RankScores."""
    wt_nt = read_fasta(source_dir / "NC_002483.1_ccdB_CDS.fasta")
    wt_aa = translate_dna(wt_nt)
    workbook_path = source_dir / "Adkar_2012_CcdB_original_results_mmc2.xlsx"
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)

    try:
        worksheet = workbook["bh01_with_RankScore"]
        headers = [cell.value for cell in worksheet[2]]
        columns = {name: index for index, name in enumerate(headers)}

        for row in worksheet.iter_rows(min_row=3, values_only=True):
            mutation = str(row[columns["Mutant"]])
            match = MUTATION_PATTERN.fullmatch(mutation)

            if match is None:
                raise ValueError(f"Unsupported Adkar mutation: {mutation}")

            wt_residue, position_text, mutant_residue, mutant_codon = match.groups()
            position = int(position_text)
            mutant_nt = replace_codon(wt_nt, position, mutant_codon)
            mutant_aa = translate_dna(mutant_nt)

            if wt_aa[position - 1] != wt_residue:
                raise ValueError(f"WT residue mismatch for {mutation}")

            if mutant_aa[position - 1] != mutant_residue:
                raise ValueError(f"Mutant residue mismatch for {mutation}")

            yield {
                "panel": "evo1",
                "study_id": "adkar_2012",
                "assay_id": ASSAY_ID,
                "variant_id": mutation,
                "organism": "Escherichia coli",
                "target": "CcdB toxin",
                "wt_nt": wt_nt,
                "mutant_nt": mutant_nt,
                "nt_edit": describe_coding_edit(wt_nt, mutant_nt),
                "wt_aa": wt_aa,
                "mutant_aa": mutant_aa,
                "aa_change": f"{wt_residue}{position}{mutant_residue}",
                "experimental_score": str(row[columns["RankScore"]]),
                "directionality": -1,
            }
    finally:
        workbook.close()


def standardize(source_dir: Path, output_dir: Path) -> int:
    """Write the standardized Adkar assay."""
    return write_variants(
        iter_variants(source_dir),
        output_dir / f"{ASSAY_ID}.csv",
    )
