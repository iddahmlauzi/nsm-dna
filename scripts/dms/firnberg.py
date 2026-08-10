from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

from dms.shared import (
    VariantRecord,
    describe_coding_edit,
    replace_codon,
    translate_dna,
    write_variants,
)

ASSAY_ID = "BLAT_ECOLX_Firnberg_2014"
WORKBOOK_NAME = "supp_msu081_Data_S1-S4.xlsx"


def read_source_rows(source_dir: Path) -> tuple[str, list[tuple[object, ...]]]:
    """Read codon fitness rows and reconstruct the author WT construct."""
    workbook = load_workbook(
        source_dir / WORKBOOK_NAME,
        read_only=True,
        data_only=True,
    )

    try:
        worksheet = workbook["S1 Codon fitnesses"]
        rows = list(worksheet.iter_rows(min_row=3, values_only=True))
    finally:
        workbook.close()

    wt_codons: dict[object, str] = {}

    for row in rows:
        position_label = row[0]
        wt_codon = str(row[1]).upper()

        if position_label not in wt_codons:
            wt_codons[position_label] = wt_codon
        elif wt_codons[position_label] != wt_codon:
            raise ValueError(f"Inconsistent WT codon at {position_label}")

    return "".join(wt_codons.values()), rows


def iter_variants(source_dir: Path) -> Iterator[VariantRecord]:
    """Convert scored TEM-1 codons from the author supplement."""
    wt_nt, rows = read_source_rows(source_dir)
    wt_aa = translate_dna(wt_nt)
    sequence_positions: dict[object, int] = {}

    for row in rows:
        position_label = row[0]
        sequence_position = sequence_positions.setdefault(
            position_label,
            len(sequence_positions) + 1,
        )
        wt_codon = str(row[1]).upper()
        mutant_codon = str(row[2]).upper()
        score = row[20]

        if mutant_codon == wt_codon or score is None:
            continue

        mutant_nt = replace_codon(wt_nt, sequence_position, mutant_codon)
        mutant_aa = translate_dna(mutant_nt)
        wt_residue = wt_aa[sequence_position - 1]
        mutant_residue = mutant_aa[sequence_position - 1]

        yield {
            "panel": "evo1",
            "study_id": "firnberg_2014",
            "assay_id": ASSAY_ID,
            "variant_id": f"ambler_{position_label}_{mutant_codon}",
            "organism": "Escherichia coli",
            "target": "TEM-1 beta-lactamase",
            "wt_nt": wt_nt,
            "mutant_nt": mutant_nt,
            "nt_edit": describe_coding_edit(wt_nt, mutant_nt),
            "wt_aa": wt_aa,
            "mutant_aa": mutant_aa,
            "aa_change": (f"{wt_residue}{sequence_position}{mutant_residue}"),
            "experimental_score": str(score),
            "directionality": 1,
        }


def standardize(source_dir: Path, output_dir: Path) -> int:
    """Write the standardized Firnberg assay."""
    return write_variants(
        iter_variants(source_dir),
        output_dir / f"{ASSAY_ID}.csv",
    )
