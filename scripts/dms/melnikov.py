import io
import re
import zipfile
from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

from dms.shared import (
    CODON_TABLE,
    VariantRecord,
    describe_coding_edit,
    replace_codon,
    translate_dna,
    write_variants,
)

ASSAY_ID = "KKA2_KLEPN_Melnikov_2014"
OLIGO_PATTERN = re.compile(r"^(KKA2_KLEPN_[AB]\d+)_(\d+)_(wt|[ACDEFGHIKLMNPQRSTVWY])$")
SCORE_FILES = (
    "Supplementary Data 2 - Unpacked/KKA2_S1_Kan18_L1.aadiff.txt",
    "Supplementary Data 2 - Unpacked/KKA2_S3_Kan18_L1.aadiff.txt",
)


def read_wt_sequence(source_dir: Path) -> str:
    """Read the exact synthetic APH(3')II construct."""
    workbook = load_workbook(
        source_dir / "supp_gku511_nar-01049-met-h-2014-File006.xlsx",
        read_only=True,
        data_only=True,
    )

    try:
        worksheet = workbook["Primers_etc"]

        for row in worksheet.iter_rows(values_only=True):
            if row[0] == "KKA2_KLEPN_opt":
                return str(row[2]).upper()
    finally:
        workbook.close()

    raise ValueError("KKA2_KLEPN_opt was not found in Primers_etc")


def read_oligos(source_dir: Path) -> dict[str, str]:
    """Read the six KKA2 WT tiles and 4,997 designed mutant tiles."""
    workbook = load_workbook(
        source_dir / "supp_gku511_nar-01049-met-h-2014-File006.xlsx",
        read_only=True,
        data_only=True,
    )
    oligos = {}

    try:
        for sheet_name in ("OLS_A", "OLS_B"):
            worksheet = workbook[sheet_name]

            for oligo_id, sequence in worksheet.iter_rows(
                min_row=6,
                max_col=2,
                values_only=True,
            ):
                if isinstance(oligo_id, str) and OLIGO_PATTERN.fullmatch(oligo_id):
                    oligos[oligo_id] = str(sequence).upper()
    finally:
        workbook.close()

    return oligos


def read_score_matrix(
    source_dir: Path,
) -> dict[tuple[int, str], float]:
    """Average the two valid Kan18 amino-acid enrichment matrices."""
    archive_path = source_dir / "supp_gku511_nar-01049-met-h-2014-File007.zip"
    replicate_scores: list[dict[tuple[int, str], float]] = []

    with zipfile.ZipFile(archive_path) as archive:
        for score_filename in SCORE_FILES:
            with archive.open(score_filename) as binary_file:
                lines = io.TextIOWrapper(binary_file, encoding="utf-8")
                next(lines)
                positions = [int(value) for value in next(lines).split("\t")[1:]]
                next(lines)
                scores = {}

                for line in lines:
                    values = line.strip().split("\t")

                    if not values[0].startswith("Delta-"):
                        continue

                    mutant_residue = values[0].removeprefix("Delta-")

                    for position, value in zip(positions, values[1:]):
                        scores[(position, mutant_residue)] = float(value)

                replicate_scores.append(scores)

    return {
        key: (replicate_scores[0][key] + replicate_scores[1][key]) / 2
        for key in replicate_scores[0]
    }


def find_mutant_codon(
    wt_tile: str,
    mutant_tile: str,
    expected_wt_codon: str,
    mutant_residue: str,
) -> str:
    """Locate the designed mutant codon within a matched oligo tile."""
    candidates: set[str] = set()

    for oriented_wt, oriented_mutant in (
        (wt_tile, mutant_tile),
        (reverse_complement(wt_tile), reverse_complement(mutant_tile)),
    ):
        changed_indices = [
            index
            for index, (wt_base, mutant_base) in enumerate(
                zip(oriented_wt, oriented_mutant)
            )
            if wt_base != mutant_base
        ]

        for codon_start in range(
            max(0, min(changed_indices) - 2),
            min(changed_indices) + 1,
        ):
            codon_end = codon_start + 3
            wt_codon = oriented_wt[codon_start:codon_end]
            mutant_codon = oriented_mutant[codon_start:codon_end]

            if any(
                index < codon_start or index >= codon_end for index in changed_indices
            ):
                continue

            if wt_codon != expected_wt_codon:
                continue

            if CODON_TABLE.get(mutant_codon) == mutant_residue:
                candidates.add(mutant_codon)

    if len(candidates) != 1:
        raise ValueError("Could not identify one designed mutant codon")

    return next(iter(candidates))


def reverse_complement(sequence: str) -> str:
    """Return the reverse complement of a DNA oligo."""
    return sequence.translate(str.maketrans("ACGT", "TGCA"))[::-1]


def iter_variants(source_dir: Path) -> Iterator[VariantRecord]:
    """Convert the author-designed APH(3')II nucleotide variants."""
    wt_nt = read_wt_sequence(source_dir)
    wt_aa = translate_dna(wt_nt)
    oligos = read_oligos(source_dir)
    scores = read_score_matrix(source_dir)
    wt_tiles = {
        match.group(1): sequence
        for oligo_id, sequence in oligos.items()
        if (match := OLIGO_PATTERN.fullmatch(oligo_id)) is not None
        and match.group(3) == "wt"
    }

    for oligo_id, mutant_tile in oligos.items():
        match = OLIGO_PATTERN.fullmatch(oligo_id)

        if match is None or match.group(3) == "wt":
            continue

        tile_id, position_text, mutant_residue = match.groups()
        position = int(position_text)
        wt_tile = wt_tiles[tile_id]
        wt_residue = wt_aa[position - 1]
        codon_start = (position - 1) * 3
        wt_codon = wt_nt[codon_start : codon_start + 3]
        mutant_codon = find_mutant_codon(
            wt_tile,
            mutant_tile,
            wt_codon,
            mutant_residue,
        )
        mutant_nt = replace_codon(wt_nt, position, mutant_codon)
        mutant_aa = translate_dna(mutant_nt)

        yield {
            "panel": "evo1",
            "study_id": "melnikov_2014",
            "assay_id": ASSAY_ID,
            "variant_id": oligo_id,
            "organism": "Klebsiella pneumoniae",
            "target": "APH(3')II aminoglycoside phosphotransferase",
            "wt_nt": wt_nt,
            "mutant_nt": mutant_nt,
            "nt_edit": describe_coding_edit(wt_nt, mutant_nt),
            "wt_aa": wt_aa,
            "mutant_aa": mutant_aa,
            "aa_change": f"{wt_residue}{position}{mutant_residue}",
            "experimental_score": str(scores[(position, mutant_residue)]),
            "directionality": 1,
        }


def standardize(source_dir: Path, output_dir: Path) -> int:
    """Write the standardized Melnikov assay."""
    return write_variants(
        iter_variants(source_dir),
        output_dir / f"{ASSAY_ID}.csv",
    )
