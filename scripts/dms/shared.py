import csv
from collections.abc import Iterable
from pathlib import Path
from typing import TypedDict

from openpyxl import load_workbook
from typing_extensions import NotRequired

STANDARD_COLUMNS = (
    "panel",
    "study_id",
    "assay_id",
    "variant_id",
    "organism",
    "target",
    "wt_nt",
    "mutant_nt",
    "nt_edit",
    "wt_aa",
    "mutant_aa",
    "aa_change",
    "experimental_score",
    "directionality",
)

# fmt: off
CODON_TABLE = {
    "TTT": "F", "TTC": "F", "TTA": "L", "TTG": "L",
    "TCT": "S", "TCC": "S", "TCA": "S", "TCG": "S",
    "TAT": "Y", "TAC": "Y", "TAA": "*", "TAG": "*",
    "TGT": "C", "TGC": "C", "TGA": "*", "TGG": "W",
    "CTT": "L", "CTC": "L", "CTA": "L", "CTG": "L",
    "CCT": "P", "CCC": "P", "CCA": "P", "CCG": "P",
    "CAT": "H", "CAC": "H", "CAA": "Q", "CAG": "Q",
    "CGT": "R", "CGC": "R", "CGA": "R", "CGG": "R",
    "ATT": "I", "ATC": "I", "ATA": "I", "ATG": "M",
    "ACT": "T", "ACC": "T", "ACA": "T", "ACG": "T",
    "AAT": "N", "AAC": "N", "AAA": "K", "AAG": "K",
    "AGT": "S", "AGC": "S", "AGA": "R", "AGG": "R",
    "GTT": "V", "GTC": "V", "GTA": "V", "GTG": "V",
    "GCT": "A", "GCC": "A", "GCA": "A", "GCG": "A",
    "GAT": "D", "GAC": "D", "GAA": "E", "GAG": "E",
    "GGT": "G", "GGC": "G", "GGA": "G", "GGG": "G",
}
# fmt: on


class VariantRecord(TypedDict):
    """One variant in the shared NSM-DNA DMS table."""

    panel: str
    study_id: str
    assay_id: str
    variant_id: NotRequired[str]
    organism: str
    target: str
    wt_nt: str
    mutant_nt: str
    nt_edit: str
    wt_aa: str
    mutant_aa: str
    aa_change: str
    experimental_score: str
    directionality: int


def translate_dna(sequence: str) -> str:
    """Translate coding DNA, preserving stop codons as `*`."""
    normalized_sequence = sequence.upper()

    if len(normalized_sequence) % 3 != 0:
        raise ValueError(
            f"Coding sequence length is not divisible by three: "
            f"{len(normalized_sequence)}"
        )

    amino_acids = []

    for start in range(0, len(normalized_sequence), 3):
        codon = normalized_sequence[start : start + 3]

        if codon not in CODON_TABLE:
            raise ValueError(f"Unsupported codon: {codon}")

        amino_acids.append(CODON_TABLE[codon])

    return "".join(amino_acids)


def read_fasta(fasta_path: Path) -> dict[str, str]:
    """Read DNA sequences from a FASTA file."""
    sequences: dict[str, list[str]] = {}
    record_name = ""

    with fasta_path.open(encoding="utf-8") as fasta_file:
        for line in fasta_file:
            line = line.strip()

            if line.startswith(">"):
                record_name = line[1:]
                sequences[record_name] = []
            elif line:
                sequences[record_name].append(line)

    return {
        record_name: "".join(sequence_parts).upper()
        for record_name, sequence_parts in sequences.items()
    }


def read_worksheet(
    workbook_path: Path,
    sheet_name: str,
    min_row: int = 1,
    max_row: int | None = None,
    min_col: int = 1,
    max_col: int | None = None,
) -> list[tuple[object, ...]]:
    """Read values from one Excel worksheet."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)

    try:
        worksheet = workbook[sheet_name]
        return list(
            worksheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
        )
    finally:
        workbook.close()


def replace_codon(
    sequence: str,
    one_based_position: int,
    mutant_codon: str,
) -> str:
    """Replace one codon in a coding DNA sequence."""
    codon_start = (one_based_position - 1) * 3
    codon_end = codon_start + 3
    normalized_codon = mutant_codon.upper()

    if one_based_position < 1 or codon_end > len(sequence):
        raise ValueError(f"Codon position {one_based_position} is outside the sequence")

    if normalized_codon not in CODON_TABLE:
        raise ValueError(f"Unsupported codon: {mutant_codon}")

    return sequence[:codon_start] + normalized_codon + sequence[codon_end:]


def describe_coding_edit(wt_nt: str, mutant_nt: str) -> str:
    """Describe coding substitutions using HGVS-like notation."""
    if len(mutant_nt) > len(wt_nt):
        return "insertion"

    if len(mutant_nt) < len(wt_nt):
        return "deletion"

    substitutions = [
        f"{position}{wt_base}>{mutant_base}"
        for position, (wt_base, mutant_base) in enumerate(
            zip(wt_nt.upper(), mutant_nt.upper()),
            start=1,
        )
        if wt_base != mutant_base
    ]

    if not substitutions:
        return "c.="

    if len(substitutions) == 1:
        return f"c.{substitutions[0]}"

    return f"c.[{';'.join(substitutions)}]"


def describe_amino_acid_changes(wt_aa: str, mutant_aa: str) -> str:
    """Describe every residue changed in one protein sequence."""
    if len(mutant_aa) > len(wt_aa):
        return "insertion"

    if len(mutant_aa) < len(wt_aa):
        return "deletion"

    changes = [
        f"{wt_residue}{position}{mutant_residue}"
        for position, (wt_residue, mutant_residue) in enumerate(
            zip(wt_aa, mutant_aa),
            start=1,
        )
        if wt_residue != mutant_residue
    ]
    return ":".join(changes) if changes else "p.="


def write_variants(
    variants: Iterable[VariantRecord],
    output_path: Path,
) -> int:
    """Write standardized variants and return the row count."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    num_variants = 0

    with output_path.open("w", encoding="utf-8", newline="") as output_file:
        writer = csv.DictWriter(output_file, fieldnames=STANDARD_COLUMNS)
        writer.writeheader()

        for variant in variants:
            writer.writerow(variant)
            num_variants += 1

    return num_variants
