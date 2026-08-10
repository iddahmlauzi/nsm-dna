from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

from dms.shared import (
    VariantRecord,
    describe_coding_edit,
    replace_codon,
    translate_dna,
    write_variants,
)

ASSAY_ID = "A4GRB6_PSEAI_Chen_2020"
WORKBOOK_NAME = "elife-56707-supp2-v2.xlsx"
POSITION_COLUMN = "codon position (G2 stands for Glycine 2 from the inhouse sequence)"


def read_source_rows(source_dir: Path) -> tuple[str, list[tuple[object, ...]]]:
    """Read codon fitness rows and reconstruct the VIM-2 WT sequence."""
    workbook = load_workbook(
        source_dir / WORKBOOK_NAME,
        read_only=True,
        data_only=True,
    )

    try:
        worksheet = workbook["SF2D Codon Fitness scores"]
        headers = [cell.value for cell in worksheet[2]]
        rows = list(worksheet.iter_rows(min_row=3, values_only=True))
    finally:
        workbook.close()

    columns = {name: index for index, name in enumerate(headers)}
    wt_codons: dict[int, str] = {}

    for row in rows:
        position_value = row[columns[POSITION_COLUMN]]

        if not isinstance(position_value, int):
            continue

        wt_codon = str(row[columns["wt codon"]]).upper()

        if position_value in wt_codons and wt_codons[position_value] != wt_codon:
            raise ValueError(f"Inconsistent WT codon at {position_value}")

        wt_codons[position_value] = wt_codon

    wt_nt = "".join(wt_codons[position] for position in sorted(wt_codons))
    return wt_nt, rows


def iter_variants(source_dir: Path) -> Iterator[VariantRecord]:
    """Convert VIM-2 codons scored under 128 micrograms/mL ampicillin."""
    wt_nt, rows = read_source_rows(source_dir)
    wt_aa = translate_dna(wt_nt)

    for row in rows:
        position_value = row[2]
        score = row[6]

        if not isinstance(position_value, int) or score is None:
            continue

        mutant_codon = str(row[1]).upper()
        mutant_nt = replace_codon(wt_nt, position_value, mutant_codon)
        mutant_aa = translate_dna(mutant_nt)

        yield {
            "panel": "evo1",
            "study_id": "chen_2020",
            "assay_id": ASSAY_ID,
            "variant_id": str(row[0]),
            "organism": "Pseudomonas aeruginosa",
            "target": "VIM-2 beta-lactamase",
            "wt_nt": wt_nt,
            "mutant_nt": mutant_nt,
            "nt_edit": describe_coding_edit(wt_nt, mutant_nt),
            "wt_aa": wt_aa,
            "mutant_aa": mutant_aa,
            "aa_change": (
                f"{wt_aa[position_value - 1]}{position_value}"
                f"{mutant_aa[position_value - 1]}"
            ),
            "experimental_score": str(score),
            "directionality": 1,
        }


def standardize(source_dir: Path, output_dir: Path) -> int:
    """Write the standardized Chen assay."""
    return write_variants(
        iter_variants(source_dir),
        output_dir / f"{ASSAY_ID}.csv",
    )
