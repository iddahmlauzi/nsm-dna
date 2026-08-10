import re
from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

from dms.shared import (
    CODON_TABLE,
    VariantRecord,
    describe_coding_edit,
    replace_codon,
    translate_dna,
    write_variants,
)

ASSAY_ID = "MTH3_HAEAE_RockahShmuel_2015"
MUTATION_PATTERN = re.compile(r"([A-Z])(\d+)([A-Z*])?")


def codon_distance(first: str, second: str) -> int:
    """Count nucleotide differences between two codons."""
    return sum(
        first_base != second_base for first_base, second_base in zip(first, second)
    )


def read_wt_sequence(source_dir: Path) -> str:
    """Read positions 0 through 331 of the author HaeIII construct."""
    workbook = load_workbook(
        source_dir / "pcbi.1004421.s002.xlsx",
        read_only=True,
        data_only=True,
    )

    try:
        worksheet = workbook["raw G0"]
        wt_codons = {
            int(row[0]): str(row[2]).upper()
            for row in worksheet.iter_rows(min_row=3, values_only=True)
            if isinstance(row[0], int) and 0 <= row[0] <= 331
        }
    finally:
        workbook.close()

    return "".join(wt_codons[position] for position in range(332))


def iter_source_scores(source_dir: Path) -> Iterator[tuple[str, float]]:
    """Read the G17 scores for missense, synonymous, and nonsense variants."""
    workbook = load_workbook(
        source_dir / "pcbi.1004421.s003.xlsx",
        read_only=True,
        data_only=True,
    )
    sheet_columns = {
        "Single nt nonSyn": 8,
        "Syn": 7,
        "nonSense": 7,
    }

    try:
        for sheet_name, score_column in sheet_columns.items():
            worksheet = workbook[sheet_name]

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None and row[score_column] is not None:
                    yield str(row[0]), row[score_column]
    finally:
        workbook.close()


def iter_variants(source_dir: Path) -> Iterator[VariantRecord]:
    """Expand each reported amino-acid effect to its one-nt source codons."""
    wt_nt = read_wt_sequence(source_dir)
    wt_aa = translate_dna(wt_nt)

    for mutation, score in iter_source_scores(source_dir):
        match = MUTATION_PATTERN.fullmatch(mutation)

        if match is None:
            raise ValueError(f"Unsupported Rockah-Shmuel mutation: {mutation}")

        source_residue, source_position_text, reported_mutant = match.groups()
        source_position = int(source_position_text)
        sequence_position = source_position + 1
        wt_codon = wt_nt[(sequence_position - 1) * 3 : sequence_position * 3]

        if wt_aa[sequence_position - 1] != source_residue:
            raise ValueError(f"WT residue mismatch for {mutation}")

        target_residue = reported_mutant or source_residue

        for mutant_codon, mutant_residue in CODON_TABLE.items():
            if mutant_residue != target_residue:
                continue

            if codon_distance(wt_codon, mutant_codon) != 1:
                continue

            mutant_nt = replace_codon(wt_nt, sequence_position, mutant_codon)
            mutant_aa = translate_dna(mutant_nt)

            yield {
                "panel": "evo1",
                "study_id": "rockah_shmuel_2015",
                "assay_id": ASSAY_ID,
                "variant_id": f"position_{source_position}_{mutant_codon}",
                "organism": "Haemophilus aegyptius",
                "target": "HaeIII DNA methyltransferase",
                "wt_nt": wt_nt,
                "mutant_nt": mutant_nt,
                "nt_edit": describe_coding_edit(wt_nt, mutant_nt),
                "wt_aa": wt_aa,
                "mutant_aa": mutant_aa,
                "aa_change": (f"{source_residue}{sequence_position}{target_residue}"),
                "experimental_score": str(score),
                "directionality": 1,
            }


def standardize(source_dir: Path, output_dir: Path) -> int:
    """Write the standardized Rockah-Shmuel assay."""
    return write_variants(
        iter_variants(source_dir),
        output_dir / f"{ASSAY_ID}.csv",
    )
