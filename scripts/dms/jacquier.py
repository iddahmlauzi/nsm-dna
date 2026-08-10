import re
from collections.abc import Iterator
from pathlib import Path

from openpyxl import load_workbook

from dms.shared import (
    VariantRecord,
    describe_coding_edit,
    read_fasta,
    translate_dna,
    write_variants,
)

ASSAY_ID = "BLAT_ECOLX_Jacquier_2013"
SUBSTITUTION_PATTERN = re.compile(r"([ACGT])(\d+)([ACGT])")


def apply_substitutions(wt_nt: str, mutation: str) -> str:
    """Apply the underscore-separated nucleotide changes in one source row."""
    mutant_bases = list(wt_nt)

    for substitution in mutation.split("_"):
        match = SUBSTITUTION_PATTERN.fullmatch(substitution)

        if match is None:
            raise ValueError(f"Unsupported Jacquier mutation: {mutation}")

        wt_base, position_text, mutant_base = match.groups()
        position = int(position_text)

        if mutant_bases[position - 1] != wt_base:
            raise ValueError(f"WT base mismatch for {substitution}")

        mutant_bases[position - 1] = mutant_base

    return "".join(mutant_bases)


def iter_variants(source_dir: Path) -> Iterator[VariantRecord]:
    """Convert the reported TEM-1 nucleotide variants and MIC scores."""
    wt_nt = read_fasta(source_dir / "J01749.1_TEM-1_CDS.fasta")
    wt_aa = translate_dna(wt_nt)
    workbook = load_workbook(
        source_dir / "1215206110_sd01.xlsx",
        read_only=True,
        data_only=True,
    )

    try:
        worksheet = workbook["TEM-1 DB"]
        headers = [cell.value for cell in worksheet[1]]
        columns = {name: index for index, name in enumerate(headers)}

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            mutation = str(row[columns["Mutant_nt"]])
            mutant_nt = apply_substitutions(wt_nt, mutation)
            mutant_aa = translate_dna(mutant_nt)
            residue = int(row[columns["Residue"]])
            mutant_residue = str(row[columns["Mutant_AA"]])

            if mutant_aa[residue - 1] != mutant_residue:
                raise ValueError(f"Mutant residue mismatch for {mutation}")

            yield {
                "panel": "evo1",
                "study_id": "jacquier_2013",
                "assay_id": ASSAY_ID,
                "variant_id": mutation,
                "organism": "Escherichia coli",
                "target": "TEM-1 beta-lactamase",
                "wt_nt": wt_nt,
                "mutant_nt": mutant_nt,
                "nt_edit": describe_coding_edit(wt_nt, mutant_nt),
                "wt_aa": wt_aa,
                "mutant_aa": mutant_aa,
                "aa_change": str(row[columns["Mutation"]]),
                "experimental_score": str(row[columns["MIC_Score"]]),
                "directionality": 1,
            }
    finally:
        workbook.close()


def standardize(source_dir: Path, output_dir: Path) -> int:
    """Write the standardized Jacquier assay."""
    return write_variants(
        iter_variants(source_dir),
        output_dir / f"{ASSAY_ID}.csv",
    )
